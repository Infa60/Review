#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Recreation de la Figure 2 - "Distribution of quality scores per item"
a partir des donnees du fichier Excel d'evaluation de la qualite.

Le script :
  1. lit la feuille des resultats d'evaluation ;
  2. compte, pour chaque item, le nombre d'articles notes 0, 1 ou 2 ;
  3. dessine le graphique a barres horizontales empilees ;
  4. enregistre la figure (PNG + PDF).

Usage :
    python generate_figure2.py
    python generate_figure2.py --input mon_fichier.xlsx --output figure2.png
    python generate_figure2.py --sheet Quality_assessment_results

Dependances : matplotlib, openpyxl  (pip install matplotlib openpyxl)
"""

import argparse
import os
import textwrap
from collections import Counter

import matplotlib
matplotlib.use("Agg")  # backend sans affichage (a definir AVANT pyplot)
import matplotlib.pyplot as plt
import openpyxl

# --- Parametres d'apparence (modifiables librement) -------------------------
DEFAULT_INPUT = r"C:\Users\bourgema\OneDrive - Université de Genève\Documents\ENABLE\Review\Manuscrit_Systematic_Review\Table\Additional_2_File_quality_assessment.xlsx"
DEFAULT_SHEET = "Quality_assessment_results"   # feuille contenant les scores
DEFAULT_OUTPUT = r"C:\Users\bourgema\OneDrive - Université de Genève\Documents\ENABLE\Review\Plot\Quality_assessment_plot.png"

SCORE_COLORS = {0: "red", 1: "orange", 2: "green"}
SCORE_LABELS = {
    0: "0 - Inadequate",
    1: "1 - Partial",
    2: "2 - Adequate",
}
TITLE = "Distribution of quality scores per item"
XLABEL = "Number of articles"
LEGEND_TITLE = "Score"
WRAP_WIDTH = 23    # largeur d'enroulement des libelles d'items (en caracteres)
BAR_HEIGHT = 0.65  # epaisseur des barres


def load_counts(path, sheet_name):
    """Renvoie (items, counts) ou counts[i] = {0: n0, 1: n1, 2: n2}."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(
            f"Feuille '{sheet_name}' introuvable. "
            f"Feuilles disponibles : {wb.sheetnames}"
        )
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]

    # Repere automatiquement les colonnes d'items : celles situees
    # entre la colonne "title" et la colonne "Total".
    start = header.index("title") + 1 if "title" in header else 3
    end = header.index("Total") if "Total" in header else len(header)
    item_cols = list(range(start, end))
    items = [str(header[c]) for c in item_cols]

    data_rows = rows[1:]
    counts = []
    for c in item_cols:
        cnt = Counter(r[c] for r in data_rows if r[c] is not None)
        counts.append({s: cnt.get(s, 0) for s in (0, 1, 2)})
    return items, counts


def make_figure(items, counts, output):
    """Construit et enregistre le graphique a barres horizontales empilees."""
    labels = [textwrap.fill(it, WRAP_WIDTH) for it in items]
    y = list(range(len(items)))

    fig, ax = plt.subplots(figsize=(10, 8))

    left = [0] * len(items)
    for score in (0, 1, 2):
        widths = [c[score] for c in counts]
        ax.barh(
            y, widths, left=left, height=BAR_HEIGHT,
            color=SCORE_COLORS[score], edgecolor="black", linewidth=1,
            label=SCORE_LABELS[score],
        )
        # Etiquette numerique blanche, centree sur chaque segment (si > 0).
        for yi, w, l in zip(y, widths, left):
            if w > 0:
                ax.text(l + w / 2, yi, str(w),
                        ha="center", va="center",
                        color="white", fontweight="bold", fontsize=9)
        left = [l + w for l, w in zip(left, widths)]

    ax.set_yticks(y)
    ax.set_yticklabels(labels)
    ax.invert_yaxis()               # l'item 1 en haut
    ax.set_xlim(left=0)
    ax.set_xlabel(XLABEL)
    ax.set_title(TITLE, fontsize=14)
    ax.legend(title=LEGEND_TITLE, loc="upper left", bbox_to_anchor=(1.01, 1))

    fig.tight_layout()
    fig.savefig(output, dpi=150, bbox_inches="tight")
    # Enregistre aussi une version vectorielle PDF a cote du PNG.
    pdf_path = os.path.splitext(output)[0] + ".pdf"
    fig.savefig(pdf_path, bbox_inches="tight")
    plt.close(fig)
    print(f"Figure enregistree : {output}")
    print(f"Version PDF         : {pdf_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Genere la figure de distribution des scores de qualite."
    )
    parser.add_argument("--input", default=DEFAULT_INPUT,
                        help="Fichier Excel source (.xlsx)")
    parser.add_argument("--sheet", default=DEFAULT_SHEET,
                        help="Nom de la feuille contenant les scores")
    parser.add_argument("--output", default=DEFAULT_OUTPUT,
                        help="Fichier image de sortie (.png)")
    args = parser.parse_args()

    items, counts = load_counts(args.input, args.sheet)

    # Petit recapitulatif dans la console.
    print(f"{len(items)} items lus depuis '{args.sheet}'.")
    for it, c in zip(items, counts):
        print(f"  {it[:45]:45s}  0={c[0]:3d}  1={c[1]:3d}  2={c[2]:3d}")

    make_figure(items, counts, args.output)


if __name__ == "__main__":
    main()